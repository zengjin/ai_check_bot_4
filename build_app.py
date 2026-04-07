import os

# 定义完善的工程结构和全量源代码
PROJECT_FILES = {
    # --- Model: 状态定义 ---
    "app/core/state.py": """from typing import TypedDict, List, Dict, Any, Optional
class WorkflowState(TypedDict):
    file_path_v1: str
    file_path_v2: str
    rules_text: str
    diff_data: List[Dict[str, Any]]
    final_prompt: str
    llm_raw_response: Dict[str, Any]
    output_path: str""",

    # --- Model: Workflow 编排 ---
    "app/core/agent.py": """from langgraph.graph import StateGraph, END
from app.core.state import WorkflowState
from app.nodes.n1_extractor import extract_node
from app.nodes.n2_builder import build_node
from app.nodes.n3_invoker import invoke_node
from app.nodes.n4_editor import edit_node

def create_workflow():
    workflow = StateGraph(WorkflowState)
    workflow.add_node("n1", extract_node)
    workflow.add_node("n2", build_node)
    workflow.add_node("n3", invoke_node)
    workflow.add_node("n4", edit_node)
    workflow.set_entry_point("n1")
    workflow.add_edge("n1", "n2")
    workflow.add_edge("n2", "n3")
    workflow.add_edge("n3", "n4")
    workflow.add_edge("n4", END)
    return workflow.compile()""",

    "app/nodes/n1_extractor.py": """import pandas as pd
def extract_node(state):
    df2 = pd.read_excel(state['file_path_v2'])
    data_to_verify = df2.to_dict('records')
    return {"diff_data": data_to_verify}""",

    "app/nodes/n2_builder.py": """import json
def build_node(state):
    rules = state['rules_text']
    data_json = json.dumps(state['diff_data'], ensure_ascii=False, indent=2)
    prompt = f\"\"\"你是一个专业的数据审计专家。请根据提供的 [校验规则] 对 [数据列表] 进行审核。
    
[校验规则]
{rules}

[数据列表]
{data_json}

[输出要求]
必须返回纯 JSON 格式，不得包含任何 Markdown 格式。
Key 为数据中的 ID (或第一列的值)，Value 为对象 {{"is_error": bool, "reason": "错误描述"}}。\"\"\"
    return {"final_prompt": prompt}""",

    "app/nodes/n3_invoker.py": """import json
import os
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage

def invoke_node(state):
    llm = ChatGoogleGenerativeAI(
        model="gemini-1.5-flash", 
        temperature=0,
        google_api_key=os.getenv("GOOGLE_API_KEY")
    )
    response = llm.invoke([HumanMessage(content=state['final_prompt'])])
    content = response.content.strip()
    if content.startswith("```json"):
        content = content.split("```json")[1].split("```")[0].strip()
    elif content.startswith("```"):
        content = content.split("```")[1].split("```")[0].strip()
    return {"llm_raw_response": json.loads(content)}""",

    "app/nodes/n4_editor.py": """from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def edit_node(state):
    results = state['llm_raw_response']
    wb = load_workbook(state['file_path_v2'])
    ws = wb.active
    red_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
    last_col = ws.max_column + 1
    ws.cell(row=1, column=last_col).value = "AI 校验结果"

    for row_idx in range(2, ws.max_row + 1):
        row_id = str(ws.cell(row=row_idx, column=1).value)
        if row_id in results:
            item = results[row_id]
            if item.get('is_error'):
                for cell in ws[row_idx]: cell.fill = red_fill
                ws.cell(row=row_idx, column=last_col).value = item.get('reason')
    wb.save(state['output_path'])
    return {}""",

    "app/core/chatbot.py": """import os
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.messages import HumanMessage, SystemMessage

class ExcelChatBot:
    def __init__(self):
        self.llm = ChatGoogleGenerativeAI(model="gemini-1.5-flash", temperature=0.7)
    def ask(self, query: str, context: str = ""):
        messages = [
            SystemMessage(content="你是一个 Excel 校验助手。"),
            HumanMessage(content=f"上下文: {context}\\n\\n问题: {query}")
        ]
        return self.llm.invoke(messages).content""",

    "app/utils/file_io.py": """import os
SCENE_MAP = {"财务审核": "finance_audit", "人事入职": "hr_onboarding", "通用校验": "default_rules"}
CONFIG_DIR = "config"
class RuleFileManager:
    @staticmethod
    def get_scenes(): return list(SCENE_MAP.keys())
    @staticmethod
    def read_rule(display_name: str) -> str:
        fid = SCENE_MAP.get(display_name, "default_rules")
        path = os.path.join(CONFIG_DIR, f"{fid}.txt")
        if not os.path.exists(path):
            os.makedirs(CONFIG_DIR, exist_ok=True)
            return "请在规则维护页面输入规则。"
        with open(path, 'r', encoding='utf-8') as f: return f.read()
    @staticmethod
    def save_rule(display_name: str, content: str):
        fid = SCENE_MAP.get(display_name, "default_rules")
        os.makedirs(CONFIG_DIR, exist_ok=True)
        path = os.path.join(CONFIG_DIR, f"{fid}.txt")
        with open(path, 'w', encoding='utf-8') as f: f.write(content)""",

    "web/controller.py": """import os
import streamlit as st
from app.core.agent import create_workflow
from app.core.chatbot import ExcelChatBot
from app.utils.file_io import RuleFileManager

class WorkflowController:
    def __init__(self):
        self.workflow = create_workflow()
        self.rule_manager = RuleFileManager()
        self.chatbot = ExcelChatBot()
    def run_validation(self, f1, f2, scene_name):
        os.makedirs('data', exist_ok=True)
        p1, p2 = f"data/v1_{f1.name}", f"data/v2_{f2.name}"
        with open(p1, "wb") as f: f.write(f1.getbuffer())
        with open(p2, "wb") as f: f.write(f2.getbuffer())
        rules = self.rule_manager.read_rule(scene_name)
        inputs = {"file_path_v1": p1, "file_path_v2": p2, "rules_text": rules, "diff_data": [], "llm_raw_response": {}, "output_path": "data/output_marked.xlsx"}
        return self.workflow.invoke(inputs)""",

    "web/streamlit_app.py": """import streamlit as st
from web.controller import WorkflowController
from dotenv import load_dotenv
load_dotenv()
st.set_page_config(page_title="AI Assistant", layout="wide")
if "ctl" not in st.session_state: st.session_state.ctl = WorkflowController()
ctl = st.session_state.ctl
def main():
    st.sidebar.title("🛠️ 管理面板")
    scenes = ctl.rule_manager.get_scenes()
    selected_scene = st.sidebar.selectbox("当前业务场景", scenes)
    t1, t2, t3 = st.tabs(["📊 自动校验", "💬 AI 问答", "⚙️ 规则维护"])
    with t1:
        u1 = st.file_uploader("V1", type="xlsx", key="u1")
        u2 = st.file_uploader("V2", type="xlsx", key="u2")
        if st.button("🚀 开始校验") and u1 and u2:
            res = ctl.run_validation(u1, u2, selected_scene)
            st.session_state.last_res = str(res['llm_raw_response'])
            st.success("分析完成！")
            with open("data/output_marked.xlsx", "rb") as f:
                st.download_button("📥 下载结果", f, file_name="report.xlsx")
    with t2:
        if "msgs" not in st.session_state: st.session_state.msgs = []
        for m in st.session_state.msgs: st.chat_message(m['role']).write(m['content'])
        if p := st.chat_input("询问更多细节..."):
            st.session_state.msgs.append({"role": "user", "content": p})
            ans = ctl.chatbot.ask(p, st.session_state.get("last_res", ""))
            st.session_state.msgs.append({"role": "assistant", "content": ans})
            st.rerun()
    with t3:
        txt = ctl.rule_manager.read_rule(selected_scene)
        new = st.text_area("编辑规则", value=txt, height=300)
        if st.button("保存"): ctl.rule_manager.save_rule(selected_scene, new); st.success("已保存")
if __name__ == "__main__": main()""",

    ".env": "GOOGLE_API_KEY=YOUR_KEY_HERE",
    "config/finance_audit.txt": "1. 工资必须大于 5000",
}

def build():
    print("🚀 正在构建工程...")
    for path, content in PROJECT_FILES.items():
        # --- 核心修正点：检查是否需要创建目录 ---
        dir_name = os.path.dirname(path)
        if dir_name:  # 只有 dir_name 不为空字符串时才创建
            os.makedirs(dir_name, exist_ok=True)
        
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"✅ 写入: {path}")
    
    for d in ["app", "app/core", "app/nodes", "app/utils", "web"]:
        with open(os.path.join(d, "__init__.py"), "w") as f: pass
    print("\n🎉 构建成功！请在 .env 中填入 API Key 并启动 streamlit。")

if __name__ == "__main__":
    build()