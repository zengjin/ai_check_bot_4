from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def edit_node(state):
    results = state['llm_raw_response']
    wb = load_workbook(state['file_path_v2'])
    ws = wb.active
    red_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
    last_col = ws.max_column + 1
    ws.cell(row=1, column=last_col).value = "AI 校验结果"

    for row_idx in range(2, ws.max_row + 1):
        row_id = str(ws.cell(row=row_idx, column=1).value)
        if row_id in results:
            item = results[row_id]
            if item.get('is_error'):
                for cell in ws[row_idx]: cell.fill = red_fill
                ws.cell(row=row_idx, column=last_col).value = item.get('reason')
    wb.save(state['output_path'])
    return {}